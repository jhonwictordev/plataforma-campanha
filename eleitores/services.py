from __future__ import annotations

import csv
import io
import unicodedata
from datetime import date, datetime
from decimal import Decimal, InvalidOperation

from django.db.models import Q, QuerySet
from django.http import HttpResponse
from django.utils import timezone
from django.utils.dateparse import parse_date, parse_datetime
from openpyxl import Workbook, load_workbook

from auditoria.models import LogSeguranca
from liderancas.models import Lideranca

from .forms import ContatoCRMFormulario
from .models import ContatoCRM

BOOL_TRUE = {"1", "true", "sim", "s", "yes", "y", "autorizado"}
BOOL_FALSE = {"0", "false", "nao", "n", "no", "negado", ""}
DATE_FIELDS = {"data_nascimento"}
DATETIME_FIELDS = {"data_primeiro_contato", "data_ultimo_contato", "data_consentimento"}
DECIMAL_FIELDS = {"latitude", "longitude"}
BOOLEAN_FIELDS = {"consentimento_comunicacao"}
STATUS_MAP = {
    "novo_contato": ContatoCRM.EtapasFunil.NOVO_CONTATO,
    "novo contato": ContatoCRM.EtapasFunil.NOVO_CONTATO,
    "primeiro_atendimento": ContatoCRM.EtapasFunil.PRIMEIRO_ATENDIMENTO,
    "primeiro atendimento": ContatoCRM.EtapasFunil.PRIMEIRO_ATENDIMENTO,
    "em_relacionamento": ContatoCRM.EtapasFunil.EM_RELACIONAMENTO,
    "em relacionamento": ContatoCRM.EtapasFunil.EM_RELACIONAMENTO,
    "interessado": ContatoCRM.EtapasFunil.INTERESSADO,
    "apoiador": ContatoCRM.EtapasFunil.APOIADOR,
    "voluntario": ContatoCRM.EtapasFunil.VOLUNTARIO,
    "voluntario(a)": ContatoCRM.EtapasFunil.VOLUNTARIO,
    "lideranca": ContatoCRM.EtapasFunil.LIDERANCA,
    "lideranca confirmada": ContatoCRM.EtapasFunil.LIDERANCA,
    "inativo": ContatoCRM.EtapasFunil.INATIVO,
}
ALIAS_CAMPOS_IMPORTACAO = {
    "nome": "nome_completo",
    "nome_completo": "nome_completo",
    "telefone": "telefone",
    "celular": "telefone",
    "whatsapp": "whatsapp",
    "email": "email",
    "data_nascimento": "data_nascimento",
    "cidade": "cidade",
    "bairro": "bairro",
    "endereco": "endereco",
    "zona_eleitoral": "zona_eleitoral",
    "secao_eleitoral": "secao_eleitoral",
    "origem": "origem_contato",
    "origem_contato": "origem_contato",
    "status": "status_funil",
    "status_funil": "status_funil",
    "tags": "tags_texto",
    "observacoes": "observacoes",
    "data_primeiro_contato": "data_primeiro_contato",
    "data_ultimo_contato": "data_ultimo_contato",
    "proxima_acao": "proxima_acao",
    "consentimento": "consentimento_comunicacao",
    "consentimento_comunicacao": "consentimento_comunicacao",
    "canal_autorizado": "canal_autorizado",
    "data_consentimento": "data_consentimento",
    "latitude": "latitude",
    "longitude": "longitude",
}
COLUNAS_IMPORTACAO_EXEMPLO = [
    "nome_completo",
    "telefone",
    "whatsapp",
    "email",
    "cidade",
    "bairro",
    "origem_contato",
    "status_funil",
    "tags",
    "consentimento_comunicacao",
    "canal_autorizado",
    "data_consentimento",
    "lideranca_uuid",
]
COLUNAS_EXPORTACAO = [
    ("nome_completo", "Nome completo"),
    ("telefone", "Telefone"),
    ("whatsapp", "WhatsApp"),
    ("email", "E-mail"),
    ("cidade", "Cidade"),
    ("bairro", "Bairro"),
    ("zona_eleitoral", "Zona eleitoral"),
    ("secao_eleitoral", "Secao eleitoral"),
    ("origem_contato", "Origem"),
    ("responsavel_cadastro", "Responsavel"),
    ("lideranca_relacionada", "Lideranca relacionada"),
    ("status_funil", "Status"),
    ("tags", "Tags"),
    ("data_primeiro_contato", "Primeiro contato"),
    ("data_ultimo_contato", "Ultimo contato"),
    ("proxima_acao", "Proxima acao"),
    ("canal_autorizado", "Canal autorizado"),
    ("data_consentimento", "Data do consentimento"),
    ("possivel_duplicado", "Possivel duplicado"),
]


def normalizar_texto(valor) -> str:
    texto = str(valor or "").strip().lower()
    texto = unicodedata.normalize("NFKD", texto).encode("ascii", "ignore").decode("ascii")
    return texto


def normalizar_cabecalho(valor) -> str:
    texto = normalizar_texto(valor)
    return texto.replace("-", "_").replace("/", "_").replace(" ", "_")


def aplicar_filtros_contatos(queryset: QuerySet[ContatoCRM], filtros: dict) -> QuerySet[ContatoCRM]:
    busca = (filtros.get("busca") or "").strip()
    cidade = (filtros.get("cidade") or "").strip()
    bairro = (filtros.get("bairro") or "").strip()
    status = (filtros.get("status") or "").strip()
    responsavel = (filtros.get("responsavel") or "").strip()
    origem = (filtros.get("origem") or "").strip()
    tag = (filtros.get("tag") or "").strip()

    if busca:
        queryset = queryset.filter(
            Q(nome_completo__icontains=busca)
            | Q(telefone__icontains=busca)
            | Q(email__icontains=busca)
        )
    if cidade:
        queryset = queryset.filter(cidade__iexact=cidade)
    if bairro:
        queryset = queryset.filter(bairro__iexact=bairro)
    if status:
        queryset = queryset.filter(status_funil=status)
    if responsavel:
        queryset = queryset.filter(responsavel_cadastro_id=responsavel)
    if origem:
        queryset = queryset.filter(origem_contato__iexact=origem)
    if tag:
        tag_normalizada = normalizar_texto(tag)
        ids = [
            contato.pk
            for contato in queryset.only("pk", "tags")
            if tag_normalizada in {normalizar_texto(item) for item in (contato.tags or [])}
        ]
        queryset = queryset.filter(pk__in=ids)
    return queryset


def agrupar_contatos_por_status(queryset: QuerySet[ContatoCRM]):
    contatos = list(queryset)
    colunas = []
    for valor, rotulo in ContatoCRM.EtapasFunil.choices:
        itens = [contato for contato in contatos if contato.status_funil == valor]
        colunas.append(
            {
                "valor": valor,
                "rotulo": rotulo,
                "quantidade": len(itens),
                "contatos": itens,
            }
        )
    return colunas


def _registrar_log_operacao(evento: str, descricao: str, usuario, campanha, severidade: str = "info"):
    LogSeguranca.todos_objetos.create(
        campanha=campanha,
        usuario=usuario,
        evento=evento,
        severidade=severidade,
        descricao=descricao,
    )


def _coletar_linhas_csv(conteudo: bytes) -> list[dict]:
    for encoding in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            texto = conteudo.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    else:
        raise ValueError("Nao foi possivel decodificar o arquivo CSV informado.")

    amostra = texto[:2048]
    try:
        dialect = csv.Sniffer().sniff(amostra, delimiters=";,")
    except csv.Error:
        dialect = csv.excel
        dialect.delimiter = ";"

    leitor = csv.DictReader(io.StringIO(texto), dialect=dialect)
    if not leitor.fieldnames:
        return []
    return [linha for linha in leitor if any(str(valor or "").strip() for valor in linha.values())]


def _coletar_linhas_xlsx(conteudo: bytes) -> list[dict]:
    workbook = load_workbook(io.BytesIO(conteudo), read_only=True, data_only=True)
    worksheet = workbook.active
    linhas = list(worksheet.iter_rows(values_only=True))
    if not linhas:
        return []

    cabecalhos = [str(valor or "").strip() for valor in linhas[0]]
    retorno = []
    for linha in linhas[1:]:
        if not any(valor not in (None, "") for valor in linha):
            continue
        retorno.append({cabecalhos[indice]: valor for indice, valor in enumerate(linha)})
    return retorno


def _carregar_linhas_arquivo(arquivo) -> list[dict]:
    nome = (getattr(arquivo, "name", "") or "").lower()
    conteudo = arquivo.read()
    if nome.endswith(".csv"):
        return _coletar_linhas_csv(conteudo)
    if nome.endswith(".xlsx"):
        return _coletar_linhas_xlsx(conteudo)
    raise ValueError("Formato nao suportado. Envie um arquivo CSV ou XLSX.")


def _formatar_data(valor, incluir_hora: bool) -> str:
    if isinstance(valor, datetime):
        if timezone.is_aware(valor):
            valor = timezone.localtime(valor)
        return valor if incluir_hora else valor.date()
    if isinstance(valor, date):
        return valor

    texto = str(valor or "").strip()
    if not texto:
        return ""

    parsed_datetime = parse_datetime(texto)
    if not parsed_datetime:
        for formato in ("%d/%m/%Y %H:%M", "%d/%m/%Y %H:%M:%S", "%Y-%m-%d %H:%M", "%Y-%m-%d %H:%M:%S"):
            try:
                parsed_datetime = datetime.strptime(texto, formato)
                break
            except ValueError:
                continue
    if parsed_datetime:
        return parsed_datetime if incluir_hora else parsed_datetime.date()

    parsed_date = parse_date(texto)
    if not parsed_date:
        for formato in ("%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d"):
            try:
                parsed_date = datetime.strptime(texto, formato).date()
                break
            except ValueError:
                continue
    if parsed_date:
        return parsed_date
    return texto


def _formatar_decimal(valor):
    if valor in (None, ""):
        return ""
    if isinstance(valor, Decimal):
        return valor
    if isinstance(valor, (int, float)):
        return str(valor)
    texto = str(valor).strip().replace(",", ".")
    try:
        return Decimal(texto)
    except InvalidOperation:
        return texto


def _formatar_booleano(valor):
    if isinstance(valor, bool):
        return valor
    texto = normalizar_texto(valor)
    if texto in BOOL_TRUE:
        return True
    if texto in BOOL_FALSE:
        return False
    return valor


def _formatar_status(valor):
    texto = normalizar_texto(valor).replace("_", " ")
    return STATUS_MAP.get(texto, valor or ContatoCRM.EtapasFunil.NOVO_CONTATO)


def _resolver_lideranca(campanha, linha_normalizada: dict):
    queryset = Lideranca.objects.filter(campanha=campanha)
    referencia = linha_normalizada.get("lideranca_uuid") or linha_normalizada.get("lideranca_id")
    if referencia:
        return queryset.filter(pk=referencia).values_list("pk", flat=True).first()
    telefone = linha_normalizada.get("lideranca_telefone")
    if telefone:
        return queryset.filter(telefone=str(telefone).strip()).values_list("pk", flat=True).first()
    nome = str(linha_normalizada.get("lideranca_nome") or "").strip()
    if nome:
        lideranca = queryset.filter(Q(nome_completo__iexact=nome) | Q(nome_conhecido__iexact=nome)).first()
        if lideranca:
            return lideranca.pk
    return None


def _mapear_linha_para_formulario(linha: dict, campanha) -> dict:
    linha_normalizada = {normalizar_cabecalho(chave): valor for chave, valor in linha.items()}
    dados = {}
    for cabecalho, campo in ALIAS_CAMPOS_IMPORTACAO.items():
        if cabecalho not in linha_normalizada:
            continue
        valor = linha_normalizada[cabecalho]
        if campo in DATE_FIELDS:
            dados[campo] = _formatar_data(valor, incluir_hora=False)
        elif campo in DATETIME_FIELDS:
            dados[campo] = _formatar_data(valor, incluir_hora=True)
        elif campo in DECIMAL_FIELDS:
            dados[campo] = _formatar_decimal(valor)
        elif campo in BOOLEAN_FIELDS:
            dados[campo] = _formatar_booleano(valor)
        elif campo == "status_funil":
            dados[campo] = _formatar_status(valor)
        else:
            dados[campo] = "" if valor is None else str(valor).strip()

    if not dados.get("status_funil"):
        dados["status_funil"] = ContatoCRM.EtapasFunil.NOVO_CONTATO

    lideranca_id = _resolver_lideranca(campanha, linha_normalizada)
    if lideranca_id:
        dados["lideranca_relacionada"] = str(lideranca_id)
    return dados


def _localizar_contato_existente(campanha, telefone: str, email: str):
    queryset = ContatoCRM.objects.filter(campanha=campanha)
    telefone = str(telefone or "").strip()
    email = str(email or "").strip()
    if telefone:
        contato = queryset.filter(telefone=telefone).first()
        if contato:
            return contato
    if email:
        contato = queryset.filter(email__iexact=email).first()
        if contato:
            return contato
    return None


def importar_contatos_arquivo(*, arquivo, campanha, usuario, atualizar_existentes: bool = False) -> dict:
    linhas = _carregar_linhas_arquivo(arquivo)
    resumo = {
        "total_linhas": len(linhas),
        "criados": 0,
        "atualizados": 0,
        "ignorados": 0,
        "duplicados": 0,
        "erros": [],
    }

    for indice, linha in enumerate(linhas, start=2):
        dados_formulario = _mapear_linha_para_formulario(linha, campanha)
        contato_existente = _localizar_contato_existente(
            campanha=campanha,
            telefone=dados_formulario.get("telefone", ""),
            email=dados_formulario.get("email", ""),
        )
        if contato_existente and not atualizar_existentes:
            resumo["ignorados"] += 1
            resumo["duplicados"] += 1
            resumo["erros"].append(
                f"Linha {indice}: contato duplicado encontrado para telefone ou e-mail. Marque a atualizacao para substituir."
            )
            continue

        formulario = ContatoCRMFormulario(
            data=dados_formulario,
            instance=contato_existente,
            usuario=usuario,
        )
        if not formulario.is_valid():
            erros = "; ".join(
                f"{campo}: {', '.join(mensagens)}" for campo, mensagens in formulario.errors.items()
            )
            resumo["erros"].append(f"Linha {indice}: {erros}")
            continue

        contato = formulario.save(commit=False)
        contato.campanha = campanha
        contato.save()
        if contato_existente:
            resumo["atualizados"] += 1
        else:
            resumo["criados"] += 1

    severidade = "aviso" if resumo["erros"] else "info"
    _registrar_log_operacao(
        evento="crm_importacao_contatos",
        descricao=(
            f"Importacao de contatos concluida: {resumo['criados']} criados, "
            f"{resumo['atualizados']} atualizados, {len(resumo['erros'])} erros."
        ),
        usuario=usuario,
        campanha=campanha,
        severidade=severidade,
    )
    return resumo


def _serializar_valor_exportacao(contato: ContatoCRM, campo: str):
    valor = getattr(contato, campo)
    if campo == "status_funil":
        return contato.get_status_funil_display()
    if campo == "tags":
        return ", ".join(valor or [])
    if campo == "responsavel_cadastro":
        return str(valor) if valor else ""
    if campo == "lideranca_relacionada":
        return str(valor) if valor else ""
    if isinstance(valor, bool):
        return "Sim" if valor else "Nao"
    if isinstance(valor, datetime):
        if timezone.is_aware(valor):
            valor = timezone.localtime(valor)
        return valor.strftime("%d/%m/%Y %H:%M")
    if isinstance(valor, date):
        return valor.strftime("%d/%m/%Y")
    return "" if valor in (None, "") else str(valor)


def contatos_autorizados_para_exportacao(queryset: QuerySet[ContatoCRM]) -> QuerySet[ContatoCRM]:
    return queryset.filter(consentimento_comunicacao=True)


def exportar_contatos_csv(queryset: QuerySet[ContatoCRM], *, usuario, campanha) -> HttpResponse:
    response = HttpResponse(content_type="text/csv; charset=utf-8")
    response.write("\ufeff")
    response["Content-Disposition"] = 'attachment; filename="contatos_autorizados.csv"'
    writer = csv.writer(response, delimiter=";")
    writer.writerow([rotulo for _, rotulo in COLUNAS_EXPORTACAO])
    for contato in queryset:
        writer.writerow([_serializar_valor_exportacao(contato, campo) for campo, _ in COLUNAS_EXPORTACAO])

    _registrar_log_operacao(
        evento="crm_exportacao_contatos_csv",
        descricao=f"Exportacao CSV de contatos autorizados concluida com {queryset.count()} registros.",
        usuario=usuario,
        campanha=campanha,
    )
    return response


def exportar_contatos_xlsx(queryset: QuerySet[ContatoCRM], *, usuario, campanha) -> HttpResponse:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Contatos autorizados"
    worksheet.append([rotulo for _, rotulo in COLUNAS_EXPORTACAO])
    for contato in queryset:
        worksheet.append([_serializar_valor_exportacao(contato, campo) for campo, _ in COLUNAS_EXPORTACAO])

    buffer = io.BytesIO()
    workbook.save(buffer)
    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="contatos_autorizados.xlsx"'
    _registrar_log_operacao(
        evento="crm_exportacao_contatos_xlsx",
        descricao=f"Exportacao XLSX de contatos autorizados concluida com {queryset.count()} registros.",
        usuario=usuario,
        campanha=campanha,
    )
    return response
