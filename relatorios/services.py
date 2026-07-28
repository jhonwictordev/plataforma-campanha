from collections import defaultdict
from decimal import Decimal
from io import BytesIO
from urllib.parse import urlencode

from django.core.exceptions import PermissionDenied
from django.db.models import Count, Q, Sum
from django.http import HttpResponse
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from agenda.models import EventoAgenda
from campanhas.models import Campanha
from comunicacao.models import CampanhaComunicacao, EnvioComunicacao
from core.permissions import (
    PAPEIS_AGENDA_LEITURA,
    PAPEIS_COMUNICACAO_LEITURA,
    PAPEIS_CRM_LEITURA,
    PAPEIS_EQUIPE_LEITURA,
    PAPEIS_FINANCEIRO_LEITURA,
    PAPEIS_METAS_LEITURA,
    filtrar_queryset_por_usuario,
    usuario_tem_nivel,
)
from eleitores.models import ContatoCRM, InteracaoContato
from equipe.models import IntegranteEquipe, TarefaEquipe
from financeiro.models import LancamentoFinanceiro
from liderancas.models import InteracaoLideranca, Lideranca
from metas.models import MetaCampanha
from usuarios.models import Usuario

PAPEIS_RELATORIO_GERAL = (
    "administrador",
    "coordenador_geral",
    "coordenador_regional",
    "financeiro",
    "comunicacao",
    "mobilizador",
    "visualizador",
)


RELATORIOS_DEFINICOES = (
    {
        "chave": "contatos_cadastrados",
        "rotulo": "Contatos cadastrados",
        "descricao": "Base do CRM com cidade, bairro, status e responsavel pelo cadastro.",
        "niveis": PAPEIS_CRM_LEITURA,
        "builder": "contatos_cadastrados",
    },
    {
        "chave": "apoiadores_regiao",
        "rotulo": "Apoiadores por regiao",
        "descricao": "Consolidado territorial de apoiadores por cidade e bairro.",
        "niveis": PAPEIS_CRM_LEITURA,
        "builder": "apoiadores_regiao",
    },
    {
        "chave": "liderancas",
        "rotulo": "Liderancas",
        "descricao": "Panorama de liderancas com situacao, regiao e apoiadores estimados.",
        "niveis": PAPEIS_CRM_LEITURA,
        "builder": "liderancas",
    },
    {
        "chave": "atividades_equipe",
        "rotulo": "Atividades da equipe",
        "descricao": "Tarefas, prioridades e prazos da equipe operacional.",
        "niveis": PAPEIS_EQUIPE_LEITURA,
        "builder": "atividades_equipe",
    },
    {
        "chave": "eventos",
        "rotulo": "Eventos",
        "descricao": "Agenda filtravel por periodo, responsavel e prioridade.",
        "niveis": PAPEIS_AGENDA_LEITURA,
        "builder": "eventos",
    },
    {
        "chave": "metas",
        "rotulo": "Metas",
        "descricao": "Acompanhamento de metas por equipe, regiao e responsavel.",
        "niveis": PAPEIS_METAS_LEITURA,
        "builder": "metas",
    },
    {
        "chave": "financeiro",
        "rotulo": "Financeiro",
        "descricao": "Receitas, despesas e saldos por categoria e centro de custo.",
        "niveis": PAPEIS_FINANCEIRO_LEITURA,
        "builder": "financeiro",
    },
    {
        "chave": "comunicacoes",
        "rotulo": "Comunicacoes",
        "descricao": "Campanhas e envios autorizados com status e respostas.",
        "niveis": PAPEIS_COMUNICACAO_LEITURA,
        "builder": "comunicacoes",
    },
    {
        "chave": "historico_interacoes",
        "rotulo": "Historico de interacoes",
        "descricao": "Linha do tempo de contatos e liderancas por responsavel.",
        "niveis": PAPEIS_CRM_LEITURA,
        "builder": "historico_interacoes",
    },
    {
        "chave": "produtividade_responsavel",
        "rotulo": "Produtividade por responsavel",
        "descricao": "Consolidado operacional por usuario dentro da campanha.",
        "niveis": PAPEIS_RELATORIO_GERAL,
        "builder": "produtividade_responsavel",
    },
)


def usuario_admin(usuario) -> bool:
    return bool(usuario.is_superuser or getattr(usuario, "is_staff", False))


def obter_relatorios_disponiveis(usuario):
    return [item for item in RELATORIOS_DEFINICOES if usuario_tem_nivel(usuario, item["niveis"])]


def obter_relatorio(chave: str):
    for item in RELATORIOS_DEFINICOES:
        if item["chave"] == chave:
            return item
    return None


def validar_tipo_relatorio(usuario, tipo_relatorio, relatorios_disponiveis):
    if not tipo_relatorio:
        return
    permitidos = {item["chave"] for item in relatorios_disponiveis}
    if tipo_relatorio not in permitidos:
        raise PermissionDenied("Voce nao possui permissao para acessar este relatorio.")


def filtrar_queryset_campanha(queryset, usuario, campanha_id: str = ""):
    if usuario_admin(usuario):
        if campanha_id:
            return queryset.filter(campanha_id=campanha_id)
        return queryset
    return filtrar_queryset_por_usuario(queryset, usuario)


def campanha_atual(usuario, campanha_id: str = ""):
    if campanha_id:
        return Campanha.objects.filter(pk=campanha_id).first()
    if not usuario_admin(usuario) and getattr(usuario, "campanha_id", None):
        return Campanha.objects.filter(pk=usuario.campanha_id).first()
    return None


def queryset_responsaveis(usuario, campanha_id: str = ""):
    queryset = Usuario.objects.order_by("nome_completo", "pk")
    if usuario_admin(usuario):
        if campanha_id:
            return queryset.filter(campanha_id=campanha_id)
        return queryset
    return queryset.filter(campanha_id=getattr(usuario, "campanha_id", None))


def serialize_filters(filters: dict):
    return {
        "tipo_relatorio": filters["tipo_relatorio"],
        "campanha": filters.get("campanha_id", ""),
        "data_inicial": filters["data_inicial"].isoformat(),
        "data_final": filters["data_final"].isoformat(),
        "cidade": filters.get("cidade", ""),
        "bairro": filters.get("bairro", ""),
        "responsavel": filters.get("responsavel_id", ""),
    }


def filters_to_querystring(filters: dict):
    payload = {}
    for chave, valor in filters.items():
        if valor in ("", None):
            continue
        payload[chave] = valor.isoformat() if hasattr(valor, "isoformat") else str(valor)
    return urlencode(payload)


def format_decimal(valor):
    if valor in (None, ""):
        return "-"
    if not isinstance(valor, Decimal):
        valor = Decimal(str(valor))
    return f"R$ {valor.quantize(Decimal('0.01'))}"


def format_number(valor):
    if valor in (None, ""):
        return "-"
    return f"{valor}"


def format_date(valor):
    if not valor:
        return "-"
    return valor.strftime("%d/%m/%Y")


def format_datetime(valor):
    if not valor:
        return "-"
    return timezone.localtime(valor).strftime("%d/%m/%Y %H:%M")


def format_bool(valor):
    return "Sim" if valor else "Nao"


def _resumo_card(rotulo, valor):
    return {"rotulo": rotulo, "valor": valor}


def _apply_city_bairro(queryset, cidade: str, bairro: str, campo_cidade="cidade", campo_bairro="bairro"):
    if cidade:
        queryset = queryset.filter(**{f"{campo_cidade}__iexact": cidade})
    if bairro:
        queryset = queryset.filter(**{f"{campo_bairro}__iexact": bairro})
    return queryset


def _builder_contatos(filters, usuario):
    queryset = filtrar_queryset_campanha(
        ContatoCRM.objects.select_related("campanha", "responsavel_cadastro", "lideranca_relacionada"),
        usuario,
        filters["campanha_id"],
    )
    queryset = queryset.filter(criado_em__date__range=(filters["data_inicial"], filters["data_final"]))
    queryset = _apply_city_bairro(queryset, filters["cidade"], filters["bairro"])
    if filters["responsavel_id"]:
        queryset = queryset.filter(responsavel_cadastro_id=filters["responsavel_id"])
    queryset = queryset.order_by("nome_completo", "pk")

    linhas = []
    for contato in queryset:
        linhas.append(
            {
                "nome": contato.nome_completo,
                "cidade": contato.cidade,
                "bairro": contato.bairro or "-",
                "status": contato.get_status_funil_display(),
                "origem": contato.origem_contato or "-",
                "responsavel": contato.responsavel_cadastro.nome_completo if contato.responsavel_cadastro else "-",
                "ultimo_contato": format_datetime(contato.data_ultimo_contato),
                "consentimento": format_bool(contato.consentimento_comunicacao),
            }
        )
    return {
        "colunas": [
            {"chave": "nome", "rotulo": "Nome"},
            {"chave": "cidade", "rotulo": "Cidade"},
            {"chave": "bairro", "rotulo": "Bairro"},
            {"chave": "status", "rotulo": "Status"},
            {"chave": "origem", "rotulo": "Origem"},
            {"chave": "responsavel", "rotulo": "Responsavel"},
            {"chave": "ultimo_contato", "rotulo": "Ultimo contato"},
            {"chave": "consentimento", "rotulo": "Consentimento"},
        ],
        "linhas": linhas,
        "resumo": [
            _resumo_card("Total de contatos", queryset.count()),
            _resumo_card("Apoiadores", queryset.filter(status_funil=ContatoCRM.EtapasFunil.APOIADOR).count()),
            _resumo_card("Voluntarios", queryset.filter(status_funil=ContatoCRM.EtapasFunil.VOLUNTARIO).count()),
            _resumo_card("Com consentimento", queryset.filter(consentimento_comunicacao=True).count()),
        ],
    }


def _builder_apoiadores_regiao(filters, usuario):
    queryset = filtrar_queryset_campanha(ContatoCRM.objects.select_related("lideranca_relacionada"), usuario, filters["campanha_id"])
    queryset = queryset.filter(
        criado_em__date__range=(filters["data_inicial"], filters["data_final"]),
        status_funil=ContatoCRM.EtapasFunil.APOIADOR,
    )
    queryset = _apply_city_bairro(queryset, filters["cidade"], filters["bairro"])
    if filters["responsavel_id"]:
        queryset = queryset.filter(responsavel_cadastro_id=filters["responsavel_id"])

    agregados = (
        queryset.values("cidade", "bairro")
        .annotate(
            total=Count("id"),
            liderancas_relacionadas=Count("lideranca_relacionada", distinct=True),
        )
        .order_by("-total", "cidade", "bairro")
    )
    linhas = [
        {
            "cidade": item["cidade"],
            "bairro": item["bairro"] or "-",
            "apoiadores": item["total"],
            "liderancas": item["liderancas_relacionadas"],
        }
        for item in agregados
    ]
    total_apoiadores = sum(item["apoiadores"] for item in linhas)
    return {
        "colunas": [
            {"chave": "cidade", "rotulo": "Cidade"},
            {"chave": "bairro", "rotulo": "Bairro"},
            {"chave": "apoiadores", "rotulo": "Apoiadores"},
            {"chave": "liderancas", "rotulo": "Liderancas relacionadas"},
        ],
        "linhas": linhas,
        "resumo": [
            _resumo_card("Territorios", len(linhas)),
            _resumo_card("Apoiadores", total_apoiadores),
            _resumo_card("Com lideranca vinculada", queryset.exclude(lideranca_relacionada__isnull=True).count()),
            _resumo_card("Consentimento ativo", queryset.filter(consentimento_comunicacao=True).count()),
        ],
    }


def _builder_liderancas(filters, usuario):
    queryset = filtrar_queryset_campanha(
        Lideranca.objects.select_related("campanha", "responsavel_contato"),
        usuario,
        filters["campanha_id"],
    )
    queryset = queryset.filter(criado_em__date__range=(filters["data_inicial"], filters["data_final"]))
    queryset = _apply_city_bairro(queryset, filters["cidade"], filters["bairro"])
    if filters["responsavel_id"]:
        queryset = queryset.filter(responsavel_contato_id=filters["responsavel_id"])
    queryset = queryset.order_by("cidade", "bairro", "nome_completo", "pk")

    linhas = []
    for lideranca in queryset:
        linhas.append(
            {
                "nome": lideranca.nome_conhecido or lideranca.nome_completo,
                "tipo": lideranca.get_tipo_lideranca_display(),
                "cidade": lideranca.cidade,
                "bairro": lideranca.bairro or "-",
                "regiao": lideranca.regiao_eleitoral or "-",
                "situacao": lideranca.get_situacao_relacionamento_display(),
                "apoiadores": lideranca.apoiadores_estimados,
                "responsavel": lideranca.responsavel_contato.nome_completo if lideranca.responsavel_contato else "-",
            }
        )
    return {
        "colunas": [
            {"chave": "nome", "rotulo": "Nome"},
            {"chave": "tipo", "rotulo": "Tipo"},
            {"chave": "cidade", "rotulo": "Cidade"},
            {"chave": "bairro", "rotulo": "Bairro"},
            {"chave": "regiao", "rotulo": "Regiao"},
            {"chave": "situacao", "rotulo": "Situacao"},
            {"chave": "apoiadores", "rotulo": "Apoiadores estimados"},
            {"chave": "responsavel", "rotulo": "Responsavel"},
        ],
        "linhas": linhas,
        "resumo": [
            _resumo_card("Total de liderancas", queryset.count()),
            _resumo_card("Confirmadas", queryset.filter(situacao_relacionamento=Lideranca.Situacoes.LIDERANCA_CONFIRMADA).count()),
            _resumo_card("Apoiadores estimados", queryset.aggregate(total=Sum("apoiadores_estimados"))["total"] or 0),
            _resumo_card("Com consentimento", queryset.filter(consentimento_dados=True).count()),
        ],
    }


def _builder_atividades_equipe(filters, usuario):
    queryset = filtrar_queryset_campanha(
        TarefaEquipe.objects.select_related("campanha", "responsavel", "responsavel__usuario"),
        usuario,
        filters["campanha_id"],
    )
    queryset = queryset.filter(
        Q(prazo__date__range=(filters["data_inicial"], filters["data_final"]))
        | Q(prazo__isnull=True, criado_em__date__range=(filters["data_inicial"], filters["data_final"]))
    )
    if filters["cidade"]:
        queryset = queryset.filter(responsavel__cidade_regiao__icontains=filters["cidade"])
    if filters["bairro"]:
        queryset = queryset.filter(Q(descricao__icontains=filters["bairro"]) | Q(comentarios__icontains=filters["bairro"]))
    if filters["responsavel_id"]:
        queryset = queryset.filter(responsavel__usuario_id=filters["responsavel_id"])
    queryset = queryset.order_by("status", "prazo", "pk")

    linhas = []
    for tarefa in queryset:
        linhas.append(
            {
                "titulo": tarefa.titulo,
                "responsavel": tarefa.responsavel.nome if tarefa.responsavel else "-",
                "departamento": tarefa.responsavel.departamento if tarefa.responsavel else "-",
                "status": tarefa.get_status_display(),
                "prioridade": tarefa.prioridade.capitalize(),
                "prazo": format_datetime(tarefa.prazo),
            }
        )
    return {
        "colunas": [
            {"chave": "titulo", "rotulo": "Tarefa"},
            {"chave": "responsavel", "rotulo": "Responsavel"},
            {"chave": "departamento", "rotulo": "Departamento"},
            {"chave": "status", "rotulo": "Status"},
            {"chave": "prioridade", "rotulo": "Prioridade"},
            {"chave": "prazo", "rotulo": "Prazo"},
        ],
        "linhas": linhas,
        "resumo": [
            _resumo_card("Total de tarefas", queryset.count()),
            _resumo_card("Concluidas", queryset.filter(status=TarefaEquipe.Status.CONCLUIDO).count()),
            _resumo_card("Em andamento", queryset.filter(status=TarefaEquipe.Status.EM_ANDAMENTO).count()),
            _resumo_card("Em atraso", queryset.filter(prazo__lt=timezone.now()).exclude(status__in=[TarefaEquipe.Status.CONCLUIDO, TarefaEquipe.Status.CANCELADO]).count()),
        ],
    }


def _builder_eventos(filters, usuario):
    queryset = filtrar_queryset_campanha(
        EventoAgenda.objects.select_related("campanha", "responsavel"),
        usuario,
        filters["campanha_id"],
    )
    queryset = queryset.filter(data__range=(filters["data_inicial"], filters["data_final"]))
    if filters["cidade"]:
        queryset = queryset.filter(endereco__icontains=filters["cidade"])
    if filters["bairro"]:
        queryset = queryset.filter(endereco__icontains=filters["bairro"])
    if filters["responsavel_id"]:
        queryset = queryset.filter(responsavel_id=filters["responsavel_id"])
    queryset = queryset.order_by("data", "horario_inicial", "pk")

    linhas = []
    for evento in queryset:
        linhas.append(
            {
                "titulo": evento.titulo,
                "tipo": evento.get_tipo_display(),
                "data": format_date(evento.data),
                "hora": evento.horario_inicial.strftime("%H:%M"),
                "local": evento.endereco or "-",
                "responsavel": evento.responsavel.nome_completo if evento.responsavel else "-",
                "status": evento.get_status_display(),
                "prioridade": evento.prioridade.capitalize(),
            }
        )
    return {
        "colunas": [
            {"chave": "titulo", "rotulo": "Evento"},
            {"chave": "tipo", "rotulo": "Tipo"},
            {"chave": "data", "rotulo": "Data"},
            {"chave": "hora", "rotulo": "Hora"},
            {"chave": "local", "rotulo": "Local"},
            {"chave": "responsavel", "rotulo": "Responsavel"},
            {"chave": "status", "rotulo": "Status"},
            {"chave": "prioridade", "rotulo": "Prioridade"},
        ],
        "linhas": linhas,
        "resumo": [
            _resumo_card("Total de eventos", queryset.count()),
            _resumo_card("Confirmados", queryset.filter(status=EventoAgenda.Status.CONFIRMADO).count()),
            _resumo_card("Concluidos", queryset.filter(status=EventoAgenda.Status.CONCLUIDO).count()),
            _resumo_card("Acoes de rua", queryset.filter(tipo=EventoAgenda.Tipos.ACAO_DE_RUA).count()),
        ],
    }


def _builder_metas(filters, usuario):
    queryset = filtrar_queryset_campanha(
        MetaCampanha.objects.select_related("campanha", "responsavel", "equipe", "equipe__usuario"),
        usuario,
        filters["campanha_id"],
    )
    queryset = queryset.filter(
        data_inicial__lte=filters["data_final"],
        data_final__gte=filters["data_inicial"],
    )
    if filters["cidade"]:
        queryset = queryset.filter(regiao__icontains=filters["cidade"])
    if filters["bairro"]:
        queryset = queryset.filter(regiao__icontains=filters["bairro"])
    if filters["responsavel_id"]:
        queryset = queryset.filter(responsavel_id=filters["responsavel_id"])
    queryset = queryset.order_by("data_final", "pk")

    linhas = []
    for meta in queryset:
        linhas.append(
            {
                "nome": meta.nome,
                "tipo": meta.get_tipo_display(),
                "periodo": f"{format_date(meta.data_inicial)} a {format_date(meta.data_final)}",
                "responsavel": meta.responsavel.nome_completo if meta.responsavel else "-",
                "equipe": meta.equipe.nome if meta.equipe else "-",
                "regiao": meta.regiao or "-",
                "esperado": format_decimal(meta.valor_esperado),
                "realizado": format_decimal(meta.valor_realizado),
                "percentual": f"{meta.percentual_conclusao}%",
                "status": meta.get_status_display(),
            }
        )
    return {
        "colunas": [
            {"chave": "nome", "rotulo": "Meta"},
            {"chave": "tipo", "rotulo": "Tipo"},
            {"chave": "periodo", "rotulo": "Periodo"},
            {"chave": "responsavel", "rotulo": "Responsavel"},
            {"chave": "equipe", "rotulo": "Equipe"},
            {"chave": "regiao", "rotulo": "Regiao"},
            {"chave": "esperado", "rotulo": "Esperado"},
            {"chave": "realizado", "rotulo": "Realizado"},
            {"chave": "percentual", "rotulo": "%"},
            {"chave": "status", "rotulo": "Status"},
        ],
        "linhas": linhas,
        "resumo": [
            _resumo_card("Total de metas", queryset.count()),
            _resumo_card("Concluidas", queryset.filter(status=MetaCampanha.Status.CONCLUIDA).count()),
            _resumo_card("Em risco", queryset.filter(status=MetaCampanha.Status.EM_RISCO).count()),
            _resumo_card("Atrasadas", queryset.filter(status=MetaCampanha.Status.ATRASADA).count()),
        ],
    }


def _builder_financeiro(filters, usuario):
    queryset = filtrar_queryset_campanha(
        LancamentoFinanceiro.objects.select_related("campanha", "categoria", "parceiro", "responsavel_lancamento", "centro_custo"),
        usuario,
        filters["campanha_id"],
    )
    queryset = queryset.filter(
        Q(data_vencimento__range=(filters["data_inicial"], filters["data_final"]))
        | Q(data_pagamento__range=(filters["data_inicial"], filters["data_final"]))
        | Q(data_vencimento__isnull=True, criado_em__date__range=(filters["data_inicial"], filters["data_final"]))
    )
    if filters["responsavel_id"]:
        queryset = queryset.filter(responsavel_lancamento_id=filters["responsavel_id"])
    queryset = queryset.order_by("data_vencimento", "pk")

    linhas = []
    for lancamento in queryset:
        linhas.append(
            {
                "descricao": lancamento.descricao,
                "tipo": lancamento.get_tipo_display(),
                "categoria": lancamento.categoria.nome if lancamento.categoria else "-",
                "valor": format_decimal(lancamento.valor_reais),
                "vencimento": format_date(lancamento.data_vencimento),
                "pagamento": format_date(lancamento.data_pagamento),
                "parceiro": lancamento.parceiro.nome if lancamento.parceiro else "-",
                "centro_custo": lancamento.centro_custo.nome if lancamento.centro_custo else "-",
                "status": lancamento.get_status_display(),
                "responsavel": lancamento.responsavel_lancamento.nome_completo if lancamento.responsavel_lancamento else "-",
            }
        )

    total_receitas = queryset.filter(tipo=LancamentoFinanceiro.Tipos.RECEITA).aggregate(total=Sum("valor_reais"))["total"] or Decimal("0")
    total_despesas = queryset.filter(tipo=LancamentoFinanceiro.Tipos.DESPESA).aggregate(total=Sum("valor_reais"))["total"] or Decimal("0")
    total_doacoes = queryset.filter(tipo=LancamentoFinanceiro.Tipos.DOACAO).aggregate(total=Sum("valor_reais"))["total"] or Decimal("0")
    return {
        "colunas": [
            {"chave": "descricao", "rotulo": "Descricao"},
            {"chave": "tipo", "rotulo": "Tipo"},
            {"chave": "categoria", "rotulo": "Categoria"},
            {"chave": "valor", "rotulo": "Valor"},
            {"chave": "vencimento", "rotulo": "Vencimento"},
            {"chave": "pagamento", "rotulo": "Pagamento"},
            {"chave": "parceiro", "rotulo": "Parceiro"},
            {"chave": "centro_custo", "rotulo": "Centro de custo"},
            {"chave": "status", "rotulo": "Status"},
            {"chave": "responsavel", "rotulo": "Responsavel"},
        ],
        "linhas": linhas,
        "resumo": [
            _resumo_card("Receitas", format_decimal(total_receitas)),
            _resumo_card("Despesas", format_decimal(total_despesas)),
            _resumo_card("Doacoes", format_decimal(total_doacoes)),
            _resumo_card("Saldo", format_decimal(total_receitas + total_doacoes - total_despesas)),
        ],
        "observacao": "Validacao eleitoral e contabil final deve ser feita por profissional juridico e contabil.",
    }


def _builder_comunicacoes(filters, usuario):
    queryset = filtrar_queryset_campanha(
        EnvioComunicacao.objects.select_related(
            "campanha",
            "campanha_comunicacao",
            "contato",
            "responsavel_registro",
        ),
        usuario,
        filters["campanha_id"],
    )
    queryset = queryset.filter(
        Q(data_programada__date__range=(filters["data_inicial"], filters["data_final"]))
        | Q(data_envio__date__range=(filters["data_inicial"], filters["data_final"]))
        | Q(criado_em__date__range=(filters["data_inicial"], filters["data_final"]))
    )
    queryset = _apply_city_bairro(queryset, filters["cidade"], filters["bairro"], "contato__cidade", "contato__bairro")
    if filters["responsavel_id"]:
        queryset = queryset.filter(
            Q(campanha_comunicacao__responsavel_id=filters["responsavel_id"])
            | Q(responsavel_registro_id=filters["responsavel_id"])
        )
    queryset = queryset.order_by("-criado_em", "pk")

    linhas = []
    for envio in queryset:
        linhas.append(
            {
                "campanha": envio.campanha_comunicacao.nome if envio.campanha_comunicacao else "-",
                "canal": envio.get_canal_display(),
                "contato": envio.contato.nome_completo,
                "cidade": envio.contato.cidade,
                "bairro": envio.contato.bairro or "-",
                "status": envio.get_status_display(),
                "programado": format_datetime(envio.data_programada),
                "enviado": format_datetime(envio.data_envio),
                "responsavel": (
                    envio.campanha_comunicacao.responsavel.nome_completo
                    if envio.campanha_comunicacao and envio.campanha_comunicacao.responsavel
                    else "-"
                ),
            }
        )
    return {
        "colunas": [
            {"chave": "campanha", "rotulo": "Campanha"},
            {"chave": "canal", "rotulo": "Canal"},
            {"chave": "contato", "rotulo": "Contato"},
            {"chave": "cidade", "rotulo": "Cidade"},
            {"chave": "bairro", "rotulo": "Bairro"},
            {"chave": "status", "rotulo": "Status"},
            {"chave": "programado", "rotulo": "Programado"},
            {"chave": "enviado", "rotulo": "Enviado"},
            {"chave": "responsavel", "rotulo": "Responsavel"},
        ],
        "linhas": linhas,
        "resumo": [
            _resumo_card("Envios programados", queryset.filter(status=EnvioComunicacao.Status.PROGRAMADO).count()),
            _resumo_card("Entregues", queryset.filter(status=EnvioComunicacao.Status.ENTREGUE).count()),
            _resumo_card("Com falha", queryset.filter(status=EnvioComunicacao.Status.FALHA).count()),
            _resumo_card("Respondidos", queryset.filter(status=EnvioComunicacao.Status.RESPONDIDO).count()),
        ],
    }


def _builder_historico_interacoes(filters, usuario):
    interacoes_contato = filtrar_queryset_campanha(
        InteracaoContato.objects.select_related("campanha", "contato", "responsavel"),
        usuario,
        filters["campanha_id"],
    ).filter(data_hora__date__range=(filters["data_inicial"], filters["data_final"]))
    interacoes_lideranca = filtrar_queryset_campanha(
        InteracaoLideranca.objects.select_related("campanha", "lideranca", "responsavel"),
        usuario,
        filters["campanha_id"],
    ).filter(data_hora__date__range=(filters["data_inicial"], filters["data_final"]))

    if filters["cidade"]:
        interacoes_contato = interacoes_contato.filter(contato__cidade__iexact=filters["cidade"])
        interacoes_lideranca = interacoes_lideranca.filter(lideranca__cidade__iexact=filters["cidade"])
    if filters["bairro"]:
        interacoes_contato = interacoes_contato.filter(contato__bairro__iexact=filters["bairro"])
        interacoes_lideranca = interacoes_lideranca.filter(lideranca__bairro__iexact=filters["bairro"])
    if filters["responsavel_id"]:
        interacoes_contato = interacoes_contato.filter(responsavel_id=filters["responsavel_id"])
        interacoes_lideranca = interacoes_lideranca.filter(responsavel_id=filters["responsavel_id"])

    linhas = []
    for item in interacoes_contato:
        linhas.append(
            {
                "modulo": "Contato",
                "nome": item.contato.nome_completo,
                "cidade": item.contato.cidade,
                "bairro": item.contato.bairro or "-",
                "tipo": item.tipo,
                "data_hora": item.data_hora,
                "responsavel": item.responsavel.nome_completo if item.responsavel else "-",
                "resumo": item.descricao,
            }
        )
    for item in interacoes_lideranca:
        linhas.append(
            {
                "modulo": "Lideranca",
                "nome": item.lideranca.nome_conhecido or item.lideranca.nome_completo,
                "cidade": item.lideranca.cidade,
                "bairro": item.lideranca.bairro or "-",
                "tipo": item.get_tipo_display(),
                "data_hora": item.data_hora,
                "responsavel": item.responsavel.nome_completo if item.responsavel else "-",
                "resumo": item.resumo,
            }
        )
    linhas.sort(key=lambda item: item["data_hora"], reverse=True)
    linhas_formatadas = [
        {
            **item,
            "data_hora": format_datetime(item["data_hora"]),
        }
        for item in linhas
    ]
    return {
        "colunas": [
            {"chave": "modulo", "rotulo": "Modulo"},
            {"chave": "nome", "rotulo": "Nome"},
            {"chave": "cidade", "rotulo": "Cidade"},
            {"chave": "bairro", "rotulo": "Bairro"},
            {"chave": "tipo", "rotulo": "Tipo"},
            {"chave": "data_hora", "rotulo": "Data e hora"},
            {"chave": "responsavel", "rotulo": "Responsavel"},
            {"chave": "resumo", "rotulo": "Resumo"},
        ],
        "linhas": linhas_formatadas,
        "resumo": [
            _resumo_card("Interacoes com contatos", interacoes_contato.count()),
            _resumo_card("Interacoes com liderancas", interacoes_lideranca.count()),
            _resumo_card("Total no periodo", len(linhas_formatadas)),
            _resumo_card("Responsaveis envolvidos", len({item["responsavel"] for item in linhas_formatadas if item["responsavel"] != "-"})),
        ],
    }


def _builder_produtividade(filters, usuario):
    usuarios_queryset = queryset_responsaveis(usuario, filters["campanha_id"])
    if filters["responsavel_id"]:
        usuarios_queryset = usuarios_queryset.filter(pk=filters["responsavel_id"])

    contatos = filtrar_queryset_campanha(ContatoCRM.objects.all(), usuario, filters["campanha_id"]).filter(
        criado_em__date__range=(filters["data_inicial"], filters["data_final"])
    )
    liderancas = filtrar_queryset_campanha(Lideranca.objects.all(), usuario, filters["campanha_id"]).filter(
        criado_em__date__range=(filters["data_inicial"], filters["data_final"])
    )
    interacoes_contato = filtrar_queryset_campanha(InteracaoContato.objects.all(), usuario, filters["campanha_id"]).filter(
        data_hora__date__range=(filters["data_inicial"], filters["data_final"])
    )
    interacoes_lideranca = filtrar_queryset_campanha(InteracaoLideranca.objects.all(), usuario, filters["campanha_id"]).filter(
        data_hora__date__range=(filters["data_inicial"], filters["data_final"])
    )
    eventos = filtrar_queryset_campanha(EventoAgenda.objects.all(), usuario, filters["campanha_id"]).filter(
        data__range=(filters["data_inicial"], filters["data_final"])
    )
    tarefas = filtrar_queryset_campanha(TarefaEquipe.objects.select_related("responsavel", "responsavel__usuario"), usuario, filters["campanha_id"]).filter(
        status=TarefaEquipe.Status.CONCLUIDO,
        atualizado_em__date__range=(filters["data_inicial"], filters["data_final"]),
    )
    metas = filtrar_queryset_campanha(MetaCampanha.objects.all(), usuario, filters["campanha_id"]).filter(
        data_inicial__lte=filters["data_final"],
        data_final__gte=filters["data_inicial"],
    )
    comunicacoes = filtrar_queryset_campanha(CampanhaComunicacao.objects.all(), usuario, filters["campanha_id"]).filter(
        criado_em__date__range=(filters["data_inicial"], filters["data_final"])
    )
    financeiro = filtrar_queryset_campanha(LancamentoFinanceiro.objects.all(), usuario, filters["campanha_id"]).filter(
        criado_em__date__range=(filters["data_inicial"], filters["data_final"])
    )

    contagem = defaultdict(lambda: defaultdict(int))
    for item in contatos.values("responsavel_cadastro_id").annotate(total=Count("id")):
        if item["responsavel_cadastro_id"]:
            contagem[str(item["responsavel_cadastro_id"])]["contatos"] = item["total"]
    for item in liderancas.values("responsavel_contato_id").annotate(total=Count("id")):
        if item["responsavel_contato_id"]:
            contagem[str(item["responsavel_contato_id"])]["liderancas"] = item["total"]
    for item in interacoes_contato.values("responsavel_id").annotate(total=Count("id")):
        if item["responsavel_id"]:
            contagem[str(item["responsavel_id"])]["interacoes"] += item["total"]
    for item in interacoes_lideranca.values("responsavel_id").annotate(total=Count("id")):
        if item["responsavel_id"]:
            contagem[str(item["responsavel_id"])]["interacoes"] += item["total"]
    for item in eventos.values("responsavel_id").annotate(total=Count("id")):
        if item["responsavel_id"]:
            contagem[str(item["responsavel_id"])]["eventos"] = item["total"]
    for item in tarefas.values("responsavel__usuario_id").annotate(total=Count("id")):
        if item["responsavel__usuario_id"]:
            contagem[str(item["responsavel__usuario_id"])]["tarefas"] = item["total"]
    for item in metas.values("responsavel_id").annotate(total=Count("id")):
        if item["responsavel_id"]:
            contagem[str(item["responsavel_id"])]["metas"] = item["total"]
    for item in comunicacoes.values("responsavel_id").annotate(total=Count("id")):
        if item["responsavel_id"]:
            contagem[str(item["responsavel_id"])]["comunicacoes"] = item["total"]
    for item in financeiro.values("responsavel_lancamento_id").annotate(total=Count("id")):
        if item["responsavel_lancamento_id"]:
            contagem[str(item["responsavel_lancamento_id"])]["financeiro"] = item["total"]

    linhas = []
    for item in usuarios_queryset:
        chave = str(item.pk)
        dados = contagem[chave]
        total = sum(dados.values())
        if total == 0 and filters["responsavel_id"]:
            continue
        linhas.append(
            {
                "responsavel": item.nome_completo,
                "contatos": dados["contatos"],
                "liderancas": dados["liderancas"],
                "interacoes": dados["interacoes"],
                "eventos": dados["eventos"],
                "tarefas": dados["tarefas"],
                "metas": dados["metas"],
                "comunicacoes": dados["comunicacoes"],
                "financeiro": dados["financeiro"],
                "total": total,
            }
        )
    linhas.sort(key=lambda item: (-item["total"], item["responsavel"]))
    return {
        "colunas": [
            {"chave": "responsavel", "rotulo": "Responsavel"},
            {"chave": "contatos", "rotulo": "Contatos"},
            {"chave": "liderancas", "rotulo": "Liderancas"},
            {"chave": "interacoes", "rotulo": "Interacoes"},
            {"chave": "eventos", "rotulo": "Eventos"},
            {"chave": "tarefas", "rotulo": "Tarefas concluidas"},
            {"chave": "metas", "rotulo": "Metas"},
            {"chave": "comunicacoes", "rotulo": "Comunicacoes"},
            {"chave": "financeiro", "rotulo": "Lancamentos"},
            {"chave": "total", "rotulo": "Total"},
        ],
        "linhas": linhas,
        "resumo": [
            _resumo_card("Responsaveis no ranking", len(linhas)),
            _resumo_card("Contatos no periodo", sum(item["contatos"] for item in linhas)),
            _resumo_card("Interacoes registradas", sum(item["interacoes"] for item in linhas)),
            _resumo_card("Tarefas concluidas", sum(item["tarefas"] for item in linhas)),
        ],
    }


BUILDERS = {
    "contatos_cadastrados": _builder_contatos,
    "apoiadores_regiao": _builder_apoiadores_regiao,
    "liderancas": _builder_liderancas,
    "atividades_equipe": _builder_atividades_equipe,
    "eventos": _builder_eventos,
    "metas": _builder_metas,
    "financeiro": _builder_financeiro,
    "comunicacoes": _builder_comunicacoes,
    "historico_interacoes": _builder_historico_interacoes,
    "produtividade_responsavel": _builder_produtividade,
}


def gerar_relatorio(tipo_relatorio: str, filters: dict, usuario):
    definicao = obter_relatorio(tipo_relatorio)
    if not definicao:
        raise PermissionDenied("Relatorio informado nao existe.")

    resultado = BUILDERS[definicao["chave"]](filters, usuario)
    campanha = campanha_atual(usuario, filters.get("campanha_id", ""))
    filtros_label = [
        f"Periodo: {format_date(filters['data_inicial'])} a {format_date(filters['data_final'])}",
    ]
    if campanha:
        filtros_label.append(f"Campanha: {campanha.nome_campanha}")
    if filters.get("cidade"):
        filtros_label.append(f"Cidade: {filters['cidade']}")
    if filters.get("bairro"):
        filtros_label.append(f"Bairro: {filters['bairro']}")
    if filters.get("responsavel_id"):
        responsavel = Usuario.objects.filter(pk=filters["responsavel_id"]).first()
        if responsavel:
            filtros_label.append(f"Responsavel: {responsavel.nome_completo}")

    return {
        "tipo": definicao["chave"],
        "titulo": definicao["rotulo"],
        "descricao": definicao["descricao"],
        "colunas": resultado["colunas"],
        "linhas": resultado["linhas"],
        "resumo": resultado["resumo"],
        "observacao": resultado.get("observacao", ""),
        "filtros_label": filtros_label,
        "campanha_atual": campanha,
    }


def nome_arquivo_relatorio(tipo_relatorio: str, extensao: str):
    timestamp = timezone.localtime().strftime("%Y%m%d-%H%M")
    return f"relatorio-{tipo_relatorio}-{timestamp}.{extensao}"


def resposta_excel(relatorio):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Relatorio"

    worksheet["A1"] = relatorio["titulo"]
    worksheet["A1"].font = Font(bold=True, size=14)
    worksheet["A2"] = " | ".join(relatorio["filtros_label"])
    worksheet["A2"].alignment = Alignment(wrap_text=True)

    linha_cabecalho = 4
    for indice, coluna in enumerate(relatorio["colunas"], start=1):
        celula = worksheet.cell(row=linha_cabecalho, column=indice, value=coluna["rotulo"])
        celula.font = Font(bold=True, color="FFFFFF")
        celula.fill = PatternFill("solid", fgColor="1D4ED8")
        celula.alignment = Alignment(horizontal="center")

    for numero_linha, linha in enumerate(relatorio["linhas"], start=linha_cabecalho + 1):
        for indice, coluna in enumerate(relatorio["colunas"], start=1):
            valor = linha.get(coluna["chave"], "")
            worksheet.cell(row=numero_linha, column=indice, value=valor)

    for indice, coluna in enumerate(relatorio["colunas"], start=1):
        valores = [str(coluna["rotulo"])]
        for linha in relatorio["linhas"]:
            valores.append(str(linha.get(coluna["chave"], "")))
        largura = min(max(len(valor) for valor in valores) + 2, 38)
        worksheet.column_dimensions[get_column_letter(indice)].width = largura

    resumo_inicio = linha_cabecalho + len(relatorio["linhas"]) + 3
    worksheet.cell(row=resumo_inicio, column=1, value="Resumo").font = Font(bold=True)
    for deslocamento, card in enumerate(relatorio["resumo"], start=1):
        worksheet.cell(row=resumo_inicio + deslocamento, column=1, value=card["rotulo"])
        worksheet.cell(row=resumo_inicio + deslocamento, column=2, value=str(card["valor"]))

    if relatorio.get("observacao"):
        worksheet.cell(row=resumo_inicio + len(relatorio["resumo"]) + 2, column=1, value=relatorio["observacao"])

    buffer = BytesIO()
    workbook.save(buffer)
    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{nome_arquivo_relatorio(relatorio["tipo"], "xlsx")}"'
    return response


def resposta_pdf(relatorio):
    buffer = BytesIO()
    document = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        leftMargin=18,
        rightMargin=18,
        topMargin=18,
        bottomMargin=18,
    )
    styles = getSampleStyleSheet()
    elements = [
        Paragraph(relatorio["titulo"], styles["Heading2"]),
        Spacer(1, 6),
        Paragraph(" | ".join(relatorio["filtros_label"]), styles["BodyText"]),
        Spacer(1, 12),
    ]

    resumo_tabela = [["Indicador", "Valor"]]
    for item in relatorio["resumo"]:
        resumo_tabela.append([item["rotulo"], str(item["valor"])])
    tabela_resumo = Table(resumo_tabela, repeatRows=1)
    tabela_resumo.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1D4ED8")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#CBD5E1")),
                ("PADDING", (0, 0), (-1, -1), 5),
            ]
        )
    )
    elements.extend([tabela_resumo, Spacer(1, 12)])

    if relatorio["linhas"]:
        tabela_dados = [[coluna["rotulo"] for coluna in relatorio["colunas"]]]
        for linha in relatorio["linhas"]:
            tabela_dados.append([str(linha.get(coluna["chave"], "")) for coluna in relatorio["colunas"]])
        tabela = Table(tabela_dados, repeatRows=1)
        tabela.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0F172A")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, -1), 7),
                    ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#CBD5E1")),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("PADDING", (0, 0), (-1, -1), 4),
                ]
            )
        )
        elements.append(tabela)
    else:
        elements.append(Paragraph("Nenhum registro encontrado para os filtros atuais.", styles["BodyText"]))

    if relatorio.get("observacao"):
        elements.extend([Spacer(1, 12), Paragraph(relatorio["observacao"], styles["Italic"])])

    document.build(elements)
    response = HttpResponse(buffer.getvalue(), content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{nome_arquivo_relatorio(relatorio["tipo"], "pdf")}"'
    return response
